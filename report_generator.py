"""
Модуль для генерации отчетов в Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import List, Dict
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Класс для генерации Excel-отчетов"""
    
    @staticmethod
    def generate_employee_report(
        employee_name: str,
        connections: List[Dict],
        stats: Dict,
        period_name: str
    ) -> str:
        """
        Генерирует Excel-отчет по сотруднику
        
        Args:
            employee_name: ФИО сотрудника
            connections: Список подключений
            stats: Итоговая статистика
            period_name: Название периода
        
        Returns:
            Путь к созданному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        
        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        title_font = Font(name='Arial', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_font = Font(name='Arial', size=11, bold=True)
        
        # Заголовок отчета
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Сводный отчет по монтажнику"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Информация о сотруднике и периоде
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Исполнитель: {employee_name}"
        ws['A2'].font = Font(name='Arial', size=11, bold=True)
        ws['A2'].alignment = cell_alignment
        
        ws.merge_cells('A3:H3')
        ws['A3'] = f"Период: {period_name}"
        ws['A3'].font = Font(name='Arial', size=11)
        ws['A3'].alignment = cell_alignment
        
        ws.merge_cells('A4:H4')
        ws['A4'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A4'].font = Font(name='Arial', size=10)
        ws['A4'].alignment = cell_alignment
        
        # Заголовки столбцов (строка 6)
        headers = [
            'Столбец',
            'Исполнители',
            'Адрес подключения',
            'Модель роутера',
            'Порт',
            'Кол-во ВОЛС м',
            'Кол-во Вит.пар м',
            'Дата'
        ]
        
        ws.row_dimensions[6].height = 30
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ширина столбцов
        ws.column_dimensions['A'].width = 12  # Столбец
        ws.column_dimensions['B'].width = 25  # Исполнители
        ws.column_dimensions['C'].width = 30  # Адрес
        ws.column_dimensions['D'].width = 15  # Модель роутера
        ws.column_dimensions['E'].width = 10  # Порт
        ws.column_dimensions['F'].width = 12  # ВОЛС
        ws.column_dimensions['G'].width = 12  # Витая пара
        ws.column_dimensions['H'].width = 18  # Дата
        
        # Данные подключений
        current_row = 7
        for idx, conn in enumerate(connections, 1):
            # Форматируем дату
            try:
                created_at = datetime.fromisoformat(conn['created_at'])
                date_str = created_at.strftime('%d.%m.%Y %H:%M')
            except:
                date_str = conn['created_at']
            
            # Список исполнителей через запятую
            executors = ', '.join(conn['all_employees'])
            
            row_data = [
                idx,  # Номер по порядку
                executors,
                conn['address'],
                conn['router_model'],
                str(conn['port']),
                conn['employee_fiber_meters'],
                conn['employee_twisted_pair_meters'],
                date_str
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                
                if col_num in [6, 7]:  # Числовые столбцы
                    cell.alignment = number_alignment
                    cell.number_format = '0.00'
                else:
                    cell.alignment = cell_alignment
            
            current_row += 1
        
        # Итоги
        current_row += 1
        
        # Итого общее
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Итого общее:"
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        # Итого ВОЛС
        cell = ws.cell(row=current_row, column=6)
        cell.value = stats['total_fiber_meters']
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
        
        # Итого витая пара
        cell = ws.cell(row=current_row, column=7)
        cell.value = stats['total_twisted_pair_meters']
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
        
        # Пустая ячейка для выравнивания
        cell = ws.cell(row=current_row, column=8)
        cell.fill = total_fill
        cell.border = border
        
        # Итого для сотрудника (с учетом деления)
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"Итого {employee_name}:"
        cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        # Итого ВОЛС для сотрудника
        cell = ws.cell(row=current_row, column=6)
        cell.value = stats['total_fiber_meters']
        cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
        
        # Итого витая пара для сотрудника
        cell = ws.cell(row=current_row, column=7)
        cell.value = stats['total_twisted_pair_meters']
        cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
        
        # Пустая ячейка для выравнивания
        cell = ws.cell(row=current_row, column=8)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = border
        
        # Сохранение файла
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"report_{employee_name.replace(' ', '_')}_{timestamp}.xlsx"
        wb.save(filename)
        
        logger.info(f"Отчет создан: {filename}")
        return filename
